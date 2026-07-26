from __future__ import annotations

import argparse
import re
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook import Workbook
from PIL import Image as PILImage


TEMPLATE_HEADERS: List[str] = [
    "客户单号",
    "产品代码",
    "目的国家",
    "税号/TAX ID",
    "进口商名称",
    "BOND有效期",
    "EORI",
    "进口商地址",
    "单独报关",
    "报关方式",
    "保价服务",
    "签名服务",
    "附加服务",
    "仓库代码",
    "收件人姓名",
    "收件人公司",
    "收件人地址1",
    "收件人地址2",
    "收件人省州",
    "收件人城市",
    "收件人邮编",
    "收件人电话",
    "收件人邮箱",
    "预约链接",
    "预约码",
    "发件人姓名",
    "发件人公司",
    "发件人国家",
    "发件人省州",
    "发件人城市",
    "发件人详细地址",
    "发件人邮编",
    "发件人电话",
    "发件人邮箱",
    "是否带磁",
    "是否带电",
    "箱号",
    "Reference ID",
    "送达时段",
    "长CM",
    "宽CM",
    "高CM",
    "单箱重量KG",
    "中文申报品名",
    "英文申报品名",
    "数量",
    "数量单位",
    "单价",
    "币种",
    "净重KG",
    "毛重KG",
    "海关编码",
    "材质",
    "品牌",
    "型号",
    "用途",
    "销售链接",
    "图片",
    "备注",
]



US_STATE_NAMES = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "FL": "Florida", "GA": "Georgia", "HI": "Hawaii", "ID": "Idaho",
    "IL": "Illinois", "IN": "Indiana", "IA": "Iowa", "KS": "Kansas",
    "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi",
    "MO": "Missouri", "MT": "Montana", "NE": "Nebraska", "NV": "Nevada",
    "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico", "NY": "New York",
    "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio", "OK": "Oklahoma",
    "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah",
    "VT": "Vermont", "VA": "Virginia", "WA": "Washington", "WV": "West Virginia",
    "WI": "Wisconsin", "WY": "Wyoming", "DC": "District of Columbia",
}

def _build_merged_lookup(ws) -> Dict[Tuple[int, int], Any]:
    """Build a lookup table for all cells covered by merged ranges.

    openpyxl only stores the value in the top-left cell of a merged range.
    For data extraction we want the displayed value to be available on every
    cell inside that merged block, so we precompute a coordinate -> value map.
    """
    lookup: Dict[Tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        top_left = ws.cell(min_row, min_col).value
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                lookup[(row_idx, col_idx)] = top_left
    return lookup


def _build_merged_range_lookup(ws) -> Dict[Tuple[int, int], Tuple[int, int, int, int]]:
    """Map each cell covered by a merged range to that range's boundaries."""
    lookup: Dict[Tuple[int, int], Tuple[int, int, int, int]] = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        bounds = (min_row, min_col, max_row, max_col)
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                lookup[(row_idx, col_idx)] = bounds
    return lookup


def _sheet_cell_value(ws, row_idx: int, col_idx: int, merged_lookup: Optional[Dict[Tuple[int, int], Any]] = None) -> Any:
    value = ws.cell(row_idx, col_idx).value
    if not _is_blank(value):
        return value
    if merged_lookup is None:
        merged_lookup = _build_merged_lookup(ws)
    return merged_lookup.get((row_idx, col_idx))


def _sheet_row_values(ws, row_idx: int, max_col: int, merged_lookup: Optional[Dict[Tuple[int, int], Any]] = None) -> List[Any]:
    if merged_lookup is None:
        merged_lookup = _build_merged_lookup(ws)
    return [_sheet_cell_value(ws, row_idx, col_idx, merged_lookup) for col_idx in range(1, max_col + 1)]



def parse_us_recipient_location(address: str) -> tuple[str, str, str, str]:
    """Parse city, state code, state full name and ZIP from a US address.

    Expected tail example: ``CLARKSVILLE, TN 37040-5502 美国``.
    Returns ``(city, state_code, state_full_name, postal_code)``.
    Raises ValueError when the structure cannot be identified reliably.
    """
    value = re.sub(r"\s+", " ", str(address or "")).strip()
    # Remove a trailing Chinese/English country marker when present.
    value = re.sub(r"\s+(美国|USA|UNITED STATES)$", "", value, flags=re.IGNORECASE).strip()
    match = re.search(
        r"(?P<city>[A-Za-z][A-Za-z .'-]*?),\s*(?P<state>[A-Za-z]{2})\s+(?P<zip>\d{5}(?:-\d{4})?)$",
        value,
    )
    if not match:
        raise ValueError(f"无法从地址识别城市、省州和邮编：{address}")
    city_segment = re.sub(r"\s+", " ", match.group("city")).strip().upper()
    # The source cell contains the full street address before the city. When a
    # common street-type token is present, treat the text after its last
    # occurrence as the city (for example: GUTHRIE HWY CLARKSVILLE).
    street_types = (
        "ST", "STREET", "RD", "ROAD", "AVE", "AVENUE", "BLVD", "BOULEVARD",
        "DR", "DRIVE", "LN", "LANE", "CT", "COURT", "HWY", "HIGHWAY",
        "PKWY", "PARKWAY", "WAY", "PL", "PLACE", "TRL", "TRAIL", "CIR",
        "CIRCLE", "TER", "TERRACE",
    )
    tokens = city_segment.split()
    suffix_positions = [i for i, token in enumerate(tokens) if token.rstrip(".") in street_types]
    if suffix_positions and suffix_positions[-1] < len(tokens) - 1:
        city = " ".join(tokens[suffix_positions[-1] + 1:])
    else:
        # Conservative fallback: use the final word before the comma. The web
        # form remains editable for uncommon multi-word cities or address styles.
        city = tokens[-1] if tokens else ""
    state_code = match.group("state").upper()
    postal_code = match.group("zip")
    state_full_name = US_STATE_NAMES.get(state_code, state_code)
    return city, state_code, state_full_name, postal_code

CURRENCY_BY_COUNTRY = {
    "US": "USD",
    "USA": "USD",
    "CA": "CAD",
    "CANADA": "CAD",
    "GB": "GBP",
    "UK": "GBP",
    "DE": "EUR",
    "FR": "EUR",
    "ES": "EUR",
    "IT": "EUR",
    "NL": "EUR",
    "BE": "EUR",
    "PL": "EUR",
    "CZ": "EUR",
    "SE": "EUR",
    "DK": "EUR",
    "AU": "AUD",
}


@dataclass
class PackingRow:
    sku: str
    product_name_en: str
    product_name_cn: str
    total_cartons: Optional[float]
    qty: Optional[float]
    gw: Optional[float]
    nw: Optional[float]
    per_carton_gw: Optional[float]
    package_size: Optional[str]
    shipment_id: str
    reference_id: str
    ref_no: str
    box_group_start_row: int = 0


@dataclass
class InvoiceRow:
    sku: str
    product_name_en: str
    product_name_cn: str
    qty: Optional[float]
    unit_value: Optional[float]
    total_value: Optional[float]
    ctns: Optional[float]
    contract_no: Optional[str]
    hs_code_foreign: Optional[str]
    tax_rate: Optional[float]
    type_of_declaration: Optional[str]
    product_picture: Optional[str]
    hs_code_cn: Optional[str]
    domestic_product_name: Optional[str]
    brand: Optional[str]
    brand_type: Optional[str]
    preferential_tariff: Optional[str]
    model: Optional[str]
    material: Optional[str]
    function: Optional[str]
    voltage: Optional[str]
    capacitance: Optional[str]
    power: Optional[str]
    battery_type: Optional[str]
    supplementary_instruction: Optional[str]
    product_picture_image: Optional[bytes] = None


@dataclass
class ConversionConfig:
    product_code: str
    destination_country: str
    warehouse_code: str
    recipient_name: str
    recipient_address1: str
    recipient_address2: str = ""
    recipient_state: str = ""
    recipient_city: str = ""
    recipient_postal_code: str = ""
    recipient_phone: str = ""
    recipient_company: str = ""
    recipient_email: str = ""
    tax_id: str = ""
    importer_name: str = ""
    bond_expiry: str = ""
    eori: str = ""
    importer_address: str = ""
    report_method: str = ""
    insurance_service: str = "BJFDR"
    signature_service: str = ""
    additional_service: str = ""
    appointment_link: str = ""
    appointment_code: str = ""
    sender_name: str = ""
    sender_company: str = ""
    sender_country: str = ""
    sender_state: str = ""
    sender_city: str = ""
    sender_address: str = ""
    sender_postal_code: str = ""
    sender_phone: str = ""
    sender_email: str = ""
    delivery_time: str = ""
    use_purpose: str = "服装"
    currency: str = "USD"
    sales_link: str = "无"
    gross_weight: str = ""
    separate_customs: str = "否"
    default_magnet: str = "否"
    default_electric: str = "否"
    quantity_unit: str = "套"
    remarks: str = ""
    shipment_id_map: Dict[str, str] = field(default_factory=dict)


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _as_str(value: Any) -> Optional[str]:
    if _is_blank(value):
        return None
    return str(value).strip()


def _as_number(value: Any) -> Optional[float]:
    if _is_blank(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "")
    try:
        return float(s)
    except Exception:
        return None


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s)).upper()


def parse_package_size(text: Optional[str]) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    if not text:
        return None, None, None
    s = str(text).strip().replace("×", "*").replace("X", "*").replace("x", "*")
    parts = [p for p in s.split("*") if p.strip()]
    if len(parts) < 3:
        return None, None, None
    out: List[Optional[float]] = []
    for p in parts[:3]:
        m = re.search(r"-?\d+(?:\.\d+)?", p)
        out.append(float(m.group(0)) if m else None)
    while len(out) < 3:
        out.append(None)
    return out[0], out[1], out[2]


def require_package_size(packing_row: PackingRow) -> Tuple[float, float, float]:
    long_cm, width_cm, height_cm = parse_package_size(packing_row.package_size)
    if long_cm is None or width_cm is None or height_cm is None:
        raise ValueError(
            f"Packing List 中 SKU {packing_row.sku} 的箱规不能为空且必须可解析为长*宽*高，例如 43*33*45。"
        )
    return long_cm, width_cm, height_cm


def require_single_weight(packing_row: PackingRow) -> float:
    if packing_row.per_carton_gw is None:
        raise ValueError(
            f"Packing List 中 SKU {packing_row.sku} 的 Per Carton G.W.（KG）不能为空。"
        )
    return packing_row.per_carton_gw


def build_box_range(shipment_id: str, box_sequence: Optional[int]) -> str:
    if not shipment_id or not str(shipment_id).strip():
        raise ValueError("箱号不能为空：Shipment ID 不能为空。")
    if box_sequence is None:
        raise ValueError(f"箱号不能为空：Shipment ID {shipment_id} 未找到有效箱序号。")
    try:
        sequence = int(box_sequence)
    except Exception as exc:
        raise ValueError(f"箱号不能为空：Shipment ID {shipment_id} 的箱序号无法解析。") from exc
    if sequence <= 0:
        raise ValueError(f"箱号不能为空：Shipment ID {shipment_id} 的箱序号必须大于 0。")
    base = str(shipment_id).strip()
    if re.search(r"U\d{6}$", base):
        base = base[:-7]
    return f"{base}U{sequence:06d}"


def build_box_sequence_map(packing_rows: List[PackingRow]) -> Dict[Tuple[str, int], int]:
    """Assign a 1-based box sequence to each merged Total Cartons group.

    The Packing List is consumed top-to-bottom. Rows sharing the same shipment
    and the same merged Total Cartons block belong to the same box and reuse
    the same sequence. When the merged block changes, the box sequence advances
    by one. When the shipment changes, numbering restarts from 1.
    """
    box_sequence_map: Dict[Tuple[str, int], int] = {}
    last_sequence_by_shipment: Dict[str, int] = {}
    for row in packing_rows:
        shipment_id = (row.shipment_id or "").strip()
        if not shipment_id:
            continue
        group_key = (shipment_id, int(row.box_group_start_row))
        if group_key in box_sequence_map:
            continue
        next_sequence = last_sequence_by_shipment.get(shipment_id, 0) + 1
        box_sequence_map[group_key] = next_sequence
        last_sequence_by_shipment[shipment_id] = next_sequence
    return box_sequence_map


def build_box_group_qty_map(packing_rows: List[PackingRow]) -> Dict[Tuple[str, int], float]:
    """Sum Qty for each merged Total Cartons block.

    The net-weight logic is box-level: first sum the quantities of all SKU rows
    that belong to the same box, then divide the single-box weight by that total.
    """
    qty_map: Dict[Tuple[str, int], float] = {}
    for row in packing_rows:
        shipment_id = (row.shipment_id or "").strip()
        if not shipment_id:
            continue
        group_key = (shipment_id, int(row.box_group_start_row))
        qty_map[group_key] = qty_map.get(group_key, 0.0) + float(row.qty or 0.0)
    return qty_map

def _merge_box_group_columns(ws, packing_rows: List[PackingRow], column_indices: Sequence[int]) -> None:
    """Merge box-level columns for rows that belong to the same box.

    The worksheet body starts at row 2, and ``packing_rows`` preserves the same
    top-to-bottom order. Consecutive rows that share the same shipment ID and
    merged Total Cartons block are treated as one box group. For groups with
    more than one row, the specified columns are merged vertically so the
    workbook stores one box-level value only once.
    """
    if not packing_rows or not column_indices:
        return

    def _group_key(row: PackingRow) -> Tuple[str, int]:
        return ((row.shipment_id or "").strip(), int(row.box_group_start_row))

    start_out_row = 2
    current_key = _group_key(packing_rows[0])
    for idx, row in enumerate(packing_rows[1:], start=3):
        key = _group_key(row)
        if key != current_key:
            end_out_row = idx - 1
            if end_out_row > start_out_row:
                for col_idx in column_indices:
                    ws.merge_cells(
                        start_row=start_out_row,
                        start_column=col_idx,
                        end_row=end_out_row,
                        end_column=col_idx,
                    )
            start_out_row = idx
            current_key = key
    end_out_row = len(packing_rows) + 1
    if end_out_row > start_out_row:
        for col_idx in column_indices:
            ws.merge_cells(
                start_row=start_out_row,
                start_column=col_idx,
                end_row=end_out_row,
                end_column=col_idx,
            )


def _find_header_row(ws, required_keywords: Sequence[str], max_scan_rows: int = 20, max_scan_cols: int = 80) -> int:
    merged_lookup = _build_merged_lookup(ws)
    for row_idx in range(1, max_scan_rows + 1):
        row = _sheet_row_values(ws, row_idx, max_scan_cols, merged_lookup)
        cells = [_norm(v) for v in row if not _is_blank(v)]
        if not cells:
            continue
        if all(any(key in cell for cell in cells) for key in required_keywords):
            return row_idx
    raise ValueError(f"Could not find a header row containing {required_keywords}")


def _header_map(row: Sequence[Any]) -> Dict[str, int]:
    return {str(v).strip(): i for i, v in enumerate(row) if not _is_blank(v)}


def _iter_rows(ws, start_row: int, max_col: int, max_blank: int = 20, max_rows: int = 1200) -> Iterable[Tuple[int, List[Any]]]:
    merged_lookup = _build_merged_lookup(ws)
    blank = 0
    for row_idx in range(start_row, min(ws.max_row, max_rows) + 1):
        row = _sheet_row_values(ws, row_idx, max_col, merged_lookup)
        if all(_is_blank(v) for v in row):
            blank += 1
            if blank >= max_blank:
                break
            continue
        blank = 0
        yield row_idx, list(row)



def extract_destination_country(source_xlsx: str) -> str:
    """Extract the destination country code from INVOICE!C7.

    Expected examples include:
    - Coutry of Destination：US
    - Country of Destination: CA
    - US

    Returns an uppercase two-letter code. Raises ValueError when the cell is
    blank or cannot be parsed reliably.
    """
    wb = load_workbook(source_xlsx, data_only=True, read_only=False)
    try:
        if "INVOICE" not in wb.sheetnames:
            raise ValueError("缺少 INVOICE 工作表。")
        inv_ws = wb["INVOICE"]
        raw = _sheet_cell_value(inv_ws, 7, 3)
    finally:
        wb.close()

    if _is_blank(raw):
        raise ValueError("INVOICE!C7 为空，无法识别目的国家。")

    value = str(raw).strip().upper()
    # Prefer the text after a normal or full-width colon.
    if ":" in value or "：" in value:
        value = re.split(r"[:：]", value)[-1].strip()

    # Accept an exact country code, or the final standalone two-letter token.
    if re.fullmatch(r"[A-Z]{2}", value):
        return value
    matches = re.findall(r"(?<![A-Z])[A-Z]{2}(?![A-Z])", value)
    if matches:
        return matches[-1]
    raise ValueError(f"无法从 INVOICE!C7 识别国家二字码：{raw}")



def split_recipient_address(address: str, limit: int = 35) -> tuple[str, str]:
    """Strictly split an address by character count.

    Address 1 receives the first ``limit`` characters. Address 2 receives all
    remaining characters. No word-boundary adjustment or second-line length
    validation is performed.
    """
    value = re.sub(r"\s+", " ", str(address or "")).strip()
    return value[:limit], value[limit:]


def extract_recipient_address(source_xlsx: str) -> str:
    """Extract recipient address from INVOICE!A6.

    The current customer source format stores the delivery address directly in
    this cell. The extracted value is shown in the web form and remains editable.
    """
    wb = load_workbook(source_xlsx, data_only=True, read_only=False)
    try:
        if "INVOICE" not in wb.sheetnames:
            raise ValueError("缺少 INVOICE 工作表。")
        inv_ws = wb["INVOICE"]
        raw = _sheet_cell_value(inv_ws, 6, 1)
    finally:
        wb.close()

    if _is_blank(raw):
        raise ValueError("INVOICE!A6 为空，无法识别收件人地址。")
    value = re.sub(r"\s+", " ", str(raw)).strip()
    if not value:
        raise ValueError("INVOICE!A6 为空，无法识别收件人地址。")
    return value


def extract_recipient_phone(source_xlsx: str) -> str:
    """Extract recipient phone from INVOICE!C4.

    The source cell stores a multi-line consignee block; the phone number is
    expected on the third non-empty line. The extracted value is editable in
    the web form.
    """
    wb = load_workbook(source_xlsx, data_only=True, read_only=False)
    try:
        if "INVOICE" not in wb.sheetnames:
            raise ValueError("缺少 INVOICE 工作表。")
        inv_ws = wb["INVOICE"]
        raw = _sheet_cell_value(inv_ws, 4, 3)
    finally:
        wb.close()

    if _is_blank(raw):
        raise ValueError("INVOICE!C4 为空，无法识别收件人电话。")

    lines = [re.sub(r"\s+", " ", line).strip() for line in str(raw).splitlines() if line.strip()]
    if len(lines) < 3:
        raise ValueError("INVOICE!C4 少于三行，无法识别收件人电话。")

    third_line = lines[2]
    digits = re.sub(r"\D", "", third_line)
    if not digits:
        raise ValueError(f"INVOICE!C4 第三行未识别到数字：{third_line}")
    return digits

def _invoice_image_bytes_by_row(inv_ws) -> Dict[int, bytes]:
    """Map worksheet row number to embedded product image bytes.

    The source invoice stores one embedded image in the Product Picture column
    for most SKU rows. We associate each image with its anchor row (+1 because
    openpyxl anchors are zero-based while worksheet rows are one-based).
    """
    image_by_row: Dict[int, bytes] = {}
    for img in getattr(inv_ws, "_images", []):
        try:
            anchor = img.anchor._from
            row_num = int(anchor.row) + 1
            if row_num not in image_by_row:
                image_by_row[row_num] = img._data()
        except Exception:
            continue
    return image_by_row


def _insert_excel_image(ws, row_idx: int, col_idx: int, image_bytes: bytes, max_width: int = 110, max_height: int = 110) -> None:
    """Insert an embedded image into a worksheet cell and resize it conservatively."""
    if not image_bytes:
        return
    try:
        pil_img = PILImage.open(BytesIO(image_bytes))
        width, height = pil_img.size
        if width <= 0 or height <= 0:
            return
        scale = min(max_width / width, max_height / height, 1.0)
        excel_img = XLImage(BytesIO(image_bytes))
        excel_img.width = max(1, int(width * scale))
        excel_img.height = max(1, int(height * scale))
        ws.add_image(excel_img, f"{get_column_letter(col_idx)}{row_idx}")
        # Use a row height that can comfortably display the resized image.
        ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 0, excel_img.height * 0.75 + 6)
    except Exception:
        # Image insertion is optional: if a particular image fails to decode,
        # keep the row text values and continue.
        return



def _detect_summary_sheet(wb):
    """Detect the third worksheet by structure instead of a fixed name.

    The source file may rename the summary sheet, so we scan all non-INVOICE and
    non-Packing List sheets and pick the first one whose header row contains the
    shipment/reference columns.
    """
    excluded = {"INVOICE", "Packing List"}
    candidate_names = [name for name in wb.sheetnames if name not in excluded]
    if not candidate_names:
        raise ValueError("未找到第三张工作表。")

    for sheet_name in candidate_names:
        ws = wb[sheet_name]
        try:
            header_row = _find_header_row(ws, ["SHIPMENTID", "REFERENCEID"])
            return sheet_name, ws, header_row
        except Exception:
            continue

    # Fallback: use the first remaining sheet and try a looser match.
    sheet_name = candidate_names[0]
    ws = wb[sheet_name]
    try:
        header_row = _find_header_row(ws, ["SHIPMENTID"])
    except Exception as exc:
        raise ValueError(f"未能识别第三张工作表 {sheet_name} 的表头。") from exc
    return sheet_name, ws, header_row

def load_source_data(source_xlsx: str) -> Tuple[List[PackingRow], Dict[str, InvoiceRow], Dict[str, float], Dict[str, str]]:
    wb = load_workbook(source_xlsx, data_only=True)

    # 第三张工作表（名称可能变化）
    summary_sheet_name, sum_ws, summary_header_row = _detect_summary_sheet(wb)
    sum_merged_lookup = _build_merged_lookup(sum_ws)
    summary_counts: Dict[str, float] = {}
    summary_reference_ids: Dict[str, str] = {}
    for row_idx in range(summary_header_row + 1, sum_ws.max_row + 1):
        row = _sheet_row_values(sum_ws, row_idx, 4, sum_merged_lookup)
        if _is_blank(row[0]):
            continue
        shipment_id = str(row[0]).strip()
        summary_counts[shipment_id] = _as_number(row[3]) or 0
        reference_id = _as_str(row[1]) or ""
        summary_reference_ids[shipment_id] = reference_id

    # Packing List
    pack_ws = wb["Packing List"]
    pack_header_row = _find_header_row(pack_ws, ["SKU", "SHIPMENTID", "REFERENCEID"])
    pack_range_lookup = _build_merged_range_lookup(pack_ws)
    packing_rows: List[PackingRow] = []
    for row_idx, row in _iter_rows(pack_ws, pack_header_row + 1, 12):
        sku = _as_str(row[0])
        shipment_id = _as_str(row[9])
        reference_id = _as_str(row[10])
        ref_no = _as_str(row[11])
        if not sku or not shipment_id:
            continue
        merged_bounds = pack_range_lookup.get((row_idx, 4))
        box_group_start_row = merged_bounds[0] if merged_bounds else row_idx
        packing_rows.append(
            PackingRow(
                sku=sku,
                product_name_en=_as_str(row[1]) or "",
                product_name_cn=_as_str(row[2]) or "",
                total_cartons=_as_number(row[3]),
                qty=_as_number(row[4]),
                gw=_as_number(row[5]),
                nw=_as_number(row[6]),
                per_carton_gw=_as_number(row[7]),
                package_size=_as_str(row[8]),
                shipment_id=shipment_id,
                reference_id=reference_id or "",
                ref_no=ref_no or "",
                box_group_start_row=box_group_start_row,
            )
        )

    # Invoice
    inv_ws = wb["INVOICE"]
    inv_header_row = _find_header_row(inv_ws, ["SKU", "UNITVALUEUS$", "HSCODE（CN）"])
    invoice_rows: Dict[str, InvoiceRow] = {}
    for row_idx, row in _iter_rows(inv_ws, inv_header_row + 1, 25):
        sku = _as_str(row[0])
        if not sku:
            continue
        invoice_rows[sku] = InvoiceRow(
            sku=sku,
            product_name_en=_as_str(row[1]) or "",
            product_name_cn=_as_str(row[2]) or "",
            qty=_as_number(row[3]),
            unit_value=_as_number(row[4]),
            total_value=_as_number(row[5]),
            ctns=_as_number(row[6]),
            contract_no=_as_str(row[7]),
            hs_code_foreign=_as_str(row[8]),
            tax_rate=_as_number(row[9]),
            type_of_declaration=_as_str(row[10]),
            product_picture=_as_str(row[11]),
            hs_code_cn=_as_str(row[12]),
            domestic_product_name=_as_str(row[13]),
            brand=_as_str(row[14]),
            brand_type=_as_str(row[15]),
            preferential_tariff=_as_str(row[16]),
            model=_as_str(row[17]),
            material=_as_str(row[18]),
            function=_as_str(row[19]),
            voltage=_as_str(row[20]),
            capacitance=_as_str(row[21]),
            power=_as_str(row[22]),
            battery_type=_as_str(row[23]),
            supplementary_instruction=_as_str(row[24]),
        )

    return packing_rows, invoice_rows, summary_counts, summary_reference_ids


def _country_to_currency(country: str) -> str:
    return CURRENCY_BY_COUNTRY.get(country.strip().upper(), "") if country else ""


def generate_rows(
    packing_rows: List[PackingRow],
    invoice_rows: Dict[str, InvoiceRow],
    summary_counts: Dict[str, float],
    summary_reference_ids: Dict[str, str],
    config: ConversionConfig,
) -> List[List[Any]]:
    out_rows: List[List[Any]] = []
    box_sequence_map = build_box_sequence_map(packing_rows)
    box_qty_map = build_box_group_qty_map(packing_rows)
    for p in packing_rows:
        if not p.shipment_id.strip():
            raise ValueError("客户单号不能为空：Packing List 中存在空的 Shipment ID。")
        inv = invoice_rows.get(p.sku)
        customer_order = (config.shipment_id_map.get(p.shipment_id, p.shipment_id) or "").strip()
        if not customer_order:
            raise ValueError(f"客户单号不能为空：Shipment ID {p.shipment_id} 未映射到有效客户单号。")
        box_sequence = box_sequence_map.get((p.shipment_id.strip(), int(p.box_group_start_row)))
        if box_sequence is None:
            raise ValueError(f"箱号不能为空：Shipment ID {p.shipment_id} 的箱序号无法确定。")
        row_map: Dict[str, Any] = {h: "" for h in TEMPLATE_HEADERS}
        row_map["客户单号"] = customer_order
        row_map["产品代码"] = config.product_code
        row_map["目的国家"] = config.destination_country.upper()
        row_map["单独报关"] = config.separate_customs.strip() or "否"
        row_map["仓库代码"] = config.warehouse_code
        row_map["税号/TAX ID"] = config.tax_id
        row_map["进口商名称"] = config.importer_name
        row_map["BOND有效期"] = config.bond_expiry
        row_map["EORI"] = config.eori
        row_map["进口商地址"] = config.importer_address
        row_map["报关方式"] = config.report_method
        row_map["保价服务"] = config.insurance_service
        row_map["签名服务"] = config.signature_service
        row_map["附加服务"] = config.additional_service
        row_map["收件人姓名"] = config.recipient_name
        row_map["收件人公司"] = config.recipient_company
        row_map["收件人地址1"] = config.recipient_address1
        row_map["收件人地址2"] = config.recipient_address2
        row_map["收件人省州"] = config.recipient_state
        row_map["收件人城市"] = config.recipient_city
        row_map["收件人邮编"] = config.recipient_postal_code
        row_map["收件人电话"] = config.recipient_phone
        row_map["收件人邮箱"] = config.recipient_email
        row_map["预约链接"] = config.appointment_link
        row_map["预约码"] = config.appointment_code
        row_map["发件人姓名"] = config.sender_name
        row_map["发件人公司"] = config.sender_company
        row_map["发件人国家"] = config.sender_country
        row_map["发件人省州"] = config.sender_state
        row_map["发件人城市"] = config.sender_city
        row_map["发件人详细地址"] = config.sender_address
        row_map["发件人邮编"] = config.sender_postal_code
        row_map["发件人电话"] = config.sender_phone
        row_map["发件人邮箱"] = config.sender_email
        row_map["是否带磁"] = config.default_magnet
        row_map["是否带电"] = config.default_electric
        row_map["箱号"] = build_box_range(p.shipment_id, box_sequence)
        reference_id = (summary_reference_ids.get(p.shipment_id) or "").strip()
        if not reference_id:
            raise ValueError(f"Reference ID不能为空：第三张工作表中 Shipment ID {p.shipment_id} 未找到对应 Reference ID。")
        row_map["Reference ID"] = reference_id
        long_cm, width_cm, height_cm = require_package_size(p)
        row_map["长CM"] = long_cm
        row_map["宽CM"] = width_cm
        row_map["高CM"] = height_cm
        single_weight = require_single_weight(p)
        row_map["单箱重量KG"] = single_weight
        if not p.product_name_cn:
            raise ValueError(f"中文申报品名不能为空：SKU {p.sku} 在 Packing List 中缺少中文品名。")
        if not p.product_name_en:
            raise ValueError(f"英文申报品名不能为空：SKU {p.sku} 在 Packing List 中缺少英文品名。")
        row_map["中文申报品名"] = p.product_name_cn
        row_map["英文申报品名"] = p.product_name_en
        row_map["数量"] = p.qty
        row_map["数量单位"] = config.quantity_unit
        if inv is None or inv.unit_value is None:
            raise ValueError(f"单价不能为空：SKU {p.sku} 在 INVOICE 中缺少单价。")
        row_map["单价"] = inv.unit_value
        row_map["币种"] = config.currency.strip().upper()
        box_total_qty = box_qty_map.get((p.shipment_id.strip(), int(p.box_group_start_row)))
        if box_total_qty is None or box_total_qty <= 0:
            raise ValueError(f"净重不能为空：SKU {p.sku} 所属箱组的数量合计无效。")
        row_map["净重KG"] = round(single_weight / float(box_total_qty), 3)
        row_map["毛重KG"] = config.gross_weight.strip() if getattr(config, "gross_weight", "").strip() else ""
        if inv is None or not inv.hs_code_foreign:
            raise ValueError(f"海关编码不能为空：SKU {p.sku} 在 INVOICE 中缺少国外进口清关编码。")
        row_map["海关编码"] = inv.hs_code_foreign
        if inv is None or not inv.material:
            raise ValueError(f"材质不能为空：SKU {p.sku} 在 INVOICE 中缺少产品材质。")
        row_map["材质"] = inv.material
        if inv is None or not inv.brand:
            raise ValueError(f"品牌不能为空：SKU {p.sku} 在 INVOICE 中缺少报关品牌。")
        row_map["品牌"] = inv.brand
        if inv is None or not inv.model:
            raise ValueError(f"型号不能为空：SKU {p.sku} 在 INVOICE 中缺少报关型号。")
        row_map["型号"] = inv.model
        row_map["送达时段"] = config.delivery_time
        row_map["用途"] = config.use_purpose.strip() or "服装"
        row_map["销售链接"] = config.sales_link.strip() or "无"
        row_map["图片"] = ""
        row_map["备注"] = config.remarks
        out_rows.append([row_map[h] for h in TEMPLATE_HEADERS])
    return out_rows


def build_workbook_from_data(
    packing_rows: List[PackingRow],
    invoice_rows: Dict[str, InvoiceRow],
    summary_counts: Dict[str, float],
    summary_reference_ids: Dict[str, str],
    template_xlsx: str,
    config: ConversionConfig,
):
    if not packing_rows:
        raise ValueError("未读取到有效的 Packing List 明细，客户单号（Shipment ID）不能为空。")
    missing_customer_orders = [row.sku for row in packing_rows if not row.shipment_id.strip()]
    if missing_customer_orders:
        raise ValueError("客户单号不能为空：存在缺少 Shipment ID 的 Packing List 明细。")
    for row in packing_rows:
        mapped_customer_order = (config.shipment_id_map.get(row.shipment_id, row.shipment_id) or "").strip()
        if not mapped_customer_order:
            raise ValueError(f"客户单号不能为空：Shipment ID {row.shipment_id} 未映射到有效客户单号。")
        reference_id = (summary_reference_ids.get(row.shipment_id) or "").strip()
        if not reference_id:
            raise ValueError(f"Reference ID不能为空：第三张工作表中 Shipment ID {row.shipment_id} 未找到对应 Reference ID。")
        inv = invoice_rows.get(row.sku)
        if not row.product_name_cn:
            raise ValueError(f"中文申报品名不能为空：SKU {row.sku} 在 Packing List 中缺少中文品名。")
        if not row.product_name_en:
            raise ValueError(f"英文申报品名不能为空：SKU {row.sku} 在 Packing List 中缺少英文品名。")
    if not config.sales_link.strip():
        raise ValueError("销售链接为必填项，不能为空。")
    if not config.use_purpose.strip():
        raise ValueError("用途为必填项，不能为空。")
    if not config.currency.strip():
        raise ValueError("币种为必填项，不能为空。")
    out_rows = generate_rows(packing_rows, invoice_rows, summary_counts, summary_reference_ids, config)

    wb = load_workbook(template_xlsx)
    ws = wb["批量下单"]

    # preserve row1 header, overwrite body cells only up to the generated range
    for r_idx, row in enumerate(out_rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    box_level_columns = ["单箱重量KG", "净重KG", "毛重KG"]
    merge_col_indices = [TEMPLATE_HEADERS.index(h) + 1 for h in box_level_columns]
    _merge_box_group_columns(ws, packing_rows, merge_col_indices)

    return wb, packing_rows, out_rows, summary_counts


def build_workbook(source_xlsx: str, template_xlsx: str, config: ConversionConfig):
    packing_rows, invoice_rows, summary_counts, summary_reference_ids = load_source_data(source_xlsx)
    return build_workbook_from_data(
        packing_rows,
        invoice_rows,
        summary_counts,
        summary_reference_ids,
        template_xlsx,
        config,
    )


def save_workbook(wb: Workbook, output_xlsx: str) -> None:
    Path(output_xlsx).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def parse_args(argv: Optional[List[str]] = None):
    parser = argparse.ArgumentParser(description="Generate B2B order template from source Excel")
    parser.add_argument("source_xlsx")
    parser.add_argument("template_xlsx")
    parser.add_argument("output_xlsx")
    parser.add_argument("--product-code", required=True)
    parser.add_argument("--destination-country", required=True)
    parser.add_argument("--warehouse-code", required=True)
    parser.add_argument("--recipient-name", required=True)
    parser.add_argument("--recipient-address1", default="")
    parser.add_argument("--recipient-address2", default="")
    parser.add_argument("--recipient-phone", default="")
    parser.add_argument("--recipient-state", required=True)
    parser.add_argument("--recipient-city", required=True)
    parser.add_argument("--recipient-postal-code", required=True)
    parser.add_argument("--use-purpose", default="服装")
    parser.add_argument("--currency", default="USD")
    parser.add_argument("--sales-link", default="无")
    parser.add_argument("--gross-weight", default="")
    parser.add_argument("--quantity-unit", default="套")
    parser.add_argument("--separate-customs", default="否")
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    config = ConversionConfig(
        product_code=args.product_code,
        destination_country=args.destination_country,
        warehouse_code=args.warehouse_code,
        recipient_name=args.recipient_name,
        recipient_address1=args.recipient_address1,
        recipient_address2=args.recipient_address2,
        recipient_state=args.recipient_state,
        recipient_city=args.recipient_city,
        recipient_postal_code=args.recipient_postal_code,
        recipient_phone=args.recipient_phone,
        use_purpose=args.use_purpose,
        currency=args.currency,
        sales_link=args.sales_link,
        gross_weight=args.gross_weight,
        separate_customs=args.separate_customs,
        quantity_unit=args.quantity_unit,
    )
    wb, packing_rows, out_rows, _summary_counts = build_workbook(args.source_xlsx, args.template_xlsx, config)
    save_workbook(wb, args.output_xlsx)
    print(f"Generated {len(out_rows)} rows from {len(packing_rows)} packing rows -> {args.output_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
